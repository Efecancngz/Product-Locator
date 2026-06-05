"""
Export Service — Excel & PDF Report Generator

Provides utility methods to generate Excel (.xlsx) and PDF (.pdf) reports
from tabular data. Returns BytesIO streams for direct HTTP streaming.
"""
import io
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import os
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """Central export engine for generating Excel and PDF reports."""

    # --- Excel Generation ---

    @staticmethod
    def generate_excel(
        title: str,
        headers: List[str],
        rows: List[List],
        sheet_name: str = "Rapor"
    ) -> io.BytesIO:
        """
        Generates a styled .xlsx workbook from tabular data.

        Args:
            title: Report title displayed in the first row.
            headers: Column header names.
            rows: 2D list of cell values.
            sheet_name: Excel sheet tab name.

        Returns:
            BytesIO stream containing the .xlsx file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # --- Title Row ---
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # --- Timestamp Row ---
        ts_text = f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
        ts_cell = ws.cell(row=2, column=1, value=ts_text)
        ts_cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
        ts_cell.alignment = Alignment(horizontal="right")
        ws.row_dimensions[2].height = 18

        # --- Header Row ---
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_border = Border(
            bottom=Side(style="medium", color="1F4E79")
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = header_border
        ws.row_dimensions[3].height = 26

        # --- Data Rows ---
        alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
        data_font = Font(name="Calibri", size=10)

        for row_idx, row_data in enumerate(rows, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = Alignment(vertical="center")
                # Zebra striping
                if (row_idx - 4) % 2 == 1:
                    cell.fill = alt_fill

        # --- Auto-fit Column Widths ---
        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(headers[col_idx - 1]))
            for row_idx in range(4, len(rows) + 4):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # --- Freeze Panes ---
        ws.freeze_panes = "A4"

        # Write to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # --- PDF Generation ---

    @staticmethod
    def generate_pdf(
        title: str,
        headers: List[str],
        rows: List[List],
        orientation: str = "landscape"
    ) -> io.BytesIO:
        """
        Generates a styled PDF document with a data table.

        Args:
            title: Report title at the top of the document.
            headers: Column header names.
            rows: 2D list of cell values.
            orientation: 'landscape' or 'portrait'.

        Returns:
            BytesIO stream containing the PDF file.
        """
        output = io.BytesIO()

        page_size = landscape(A4) if orientation == "landscape" else A4

        doc = SimpleDocTemplate(
            output,
            pagesize=page_size,
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm
        )

        elements = []
        styles = getSampleStyleSheet()

        # --- Title ---
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=colors.HexColor("#1F4E79"),
            spaceAfter=6,
            alignment=1  # Center
        )
        elements.append(Paragraph(title, title_style))

        # --- Timestamp ---
        ts_text = f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ts_style = ParagraphStyle(
            "Timestamp",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.HexColor("#999999"),
            alignment=2,  # Right
            spaceAfter=12
        )
        elements.append(Paragraph(ts_text, ts_style))
        elements.append(Spacer(1, 4 * mm))

        # --- Table ---
        # Truncate long cell values for PDF readability
        def truncate(val, max_len=40):
            s = str(val) if val is not None else ""
            return s[:max_len] + "…" if len(s) > max_len else s

        table_data = [headers]
        for row in rows:
            table_data.append([truncate(cell) for cell in row])

        # Calculate column widths proportionally
        available_width = page_size[0] - 30 * mm
        num_cols = len(headers)
        col_width = available_width / num_cols if num_cols > 0 else available_width

        table = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)

        # Table styling
        table_style = TableStyle([
            # Header row
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGNMENT", (0, 0), (-1, 0), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),

            # Data rows
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),

            # Grid
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#1F4E79")),

            # Alignment
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])

        # Zebra striping
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                table_style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F2F7FB"))

        table.setStyle(table_style)
        elements.append(table)

        # --- Footer ---
        elements.append(Spacer(1, 8 * mm))
        footer_style = ParagraphStyle(
            "Footer",
            parent=styles["Normal"],
            fontSize=7,
            textColor=colors.HexColor("#AAAAAA"),
            alignment=1  # Center
        )
        elements.append(Paragraph("Product Locator — Otomatik Rapor", footer_style))

        doc.build(elements)
        output.seek(0)
        return output


# Singleton instance
export_service = ExportService()
